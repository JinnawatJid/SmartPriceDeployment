# ============================================
# quotation.py — SQLite Version (Full Feature)
# Compatible with old JSON behavior 100%
# ============================================

from __future__ import annotations
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime
import json
import sqlite3
from db_sqlite import get_conn

from fastapi import APIRouter, HTTPException, Body
from openpyxl import load_workbook

router = APIRouter(prefix="/quotation", tags=["quotation"])


# Excel template directory
DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_FILE = DATA_DIR / "QuoteTemplate.xlsx"


# -----------------------------------------------------
# Helpers
# -----------------------------------------------------
def _now_iso():
    return datetime.now().isoformat(timespec="seconds")


def _generate_quote_no(branch_code: str) -> str:
    """
    Format:  BSQO-2502/0001
    """
    now = datetime.now()
    yy = str(now.year)[-2:]
    mm = f"{now.month:02d}"

    prefix = f"{branch_code[-2:].upper()}QT-{yy}{mm}"

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT QuoteNo FROM Quote_Header
        WHERE QuoteNo LIKE ?
    """, (f"{prefix}%",))

    count = len(cur.fetchall()) + 1
    seq = f"{count:04d}"

    return f"{prefix}/{seq}"


def _safe_get_header_row(ws):
    first_row = next(ws.iter_rows(min_row=1, max_row=1))
    return [c.value for c in first_row]


def _append_header_to_excel(header: dict):
    if not EXCEL_FILE.exists():
        return

    wb = load_workbook(EXCEL_FILE)
    if "Quote_Header" not in wb.sheetnames:
        return

    ws = wb["Quote_Header"]
    headers = _safe_get_header_row(ws)
    row = [""] * len(headers)

    def set_col(col, val):
        if col in headers:
            row[headers.index(col)] = val

    # Map header values
    for key, val in header.items():
        set_col(key, val)

    ws.append(row)
    wb.save(EXCEL_FILE)


def _append_lines_to_excel(quote_no: str, lines: list):
    if not EXCEL_FILE.exists():
        return

    wb = load_workbook(EXCEL_FILE)
    if "Quote_Line" not in wb.sheetnames:
        return

    ws = wb["Quote_Line"]
    headers = _safe_get_header_row(ws)

    def to_row(line: dict):
        row = [""] * len(headers)

        def set_col(col, val):
            if col in headers:
                row[headers.index(col)] = val

        set_col("QuoteID", quote_no)
        for key, val in line.items():
            set_col(key, val)

        return row

    for ln in lines:
        ws.append(to_row(ln))

    wb.save(EXCEL_FILE)


# -----------------------------------------------------
# ⭐ Normalize keys
# -----------------------------------------------------
def normalize_keys(row: dict):
    return {k.strip(): v for k, v in row.items()}

def _build_line_from_payload(item: dict) -> dict:
    """
    สร้างโครง line ให้ DB จาก payload ของ FE
    พร้อม normalize ตัวเลข และคำนวณ lineTotal ในฝั่ง backend
    """
    sku = item["sku"]
    name = item.get("name", "")
    category = (
        item.get("category")
        or str(item.get("sku", "")).strip()[:1].upper()
    )

    unit = item.get("unit", "")

    qty = float(item.get("qty", 0) or 0)
    price = float(item.get("price", 0) or 0)

    sqft_sheet = float(item.get("sqft_sheet", 0) or 0)

    # กัน qty ติดลบโดยไม่เปลี่ยน behavior ปกติ (ถ้า FE ส่งผิด → เราก็ treat เป็น 0)
    if qty < 0:
        qty = 0

    # lineTotal คำนวณที่ backend (สูตรเดียวกับ FE → ผลลัพธ์ปกติจะเหมือนเดิม)
        # 1) ถ้า FE ส่ง lineTotal มา (จาก pricing) ให้ใช้ค่านั้นเป็น source of truth
    if item.get("lineTotal") is not None:
        line_total = float(item.get("lineTotal") or 0)

    else:
        # 2) ถ้าไม่มี lineTotal → คำนวณ fallback
        if str(category).upper() == "G":
            line_total = round(price * sqft_sheet * qty, 2)   # ✅ กระจก
        else:
            line_total = round(price * qty, 2)                # ✅ อื่นๆ


    return {
        "ItemCode": sku,
        "ItemName": name,
        "Category": category,
        "Unit": unit,
        "Quantity": qty,
        "UnitPrice": price,
        "TotalPrice": line_total,
        "IsGlassCut": "Y" if item.get("isGlassCut") else "N",
        "Sqft_Sheet": sqft_sheet,
        "VariantCode": item.get("variantCode", ""),
        "ProductWeight": float(item.get("product_weight", 0) or 0),
        "CutInfoJson": json.dumps(item.get("cutInfo", "")),
        "Remark": item.get("remark", "")
    }


# -----------------------------------------------------
# CREATE QUOTATION
# -----------------------------------------------------
@router.post("", summary="สร้างใบเสนอราคาใหม่")
def create_quotation(payload: dict = Body(...)):
    conn = get_conn()
    cur = conn.cursor()

    employee = payload.get("employee") or {}
    customer = payload.get("customer") or {}
    branch = employee.get("branchId", "")

     # -------------------------------
    #  กำหนดค่า default: ผู้ไม่ประสงค์ออกนาม
    # -------------------------------
    raw_code = (customer.get("code") or "").strip()
    raw_name = (customer.get("name") or "").strip()
    raw_phone = (customer.get("phone") or "").strip()

    if not raw_code and not raw_name:
        # ไม่กรอกอะไรเกี่ยวกับลูกค้าเลย → ถือว่าเป็นผู้ไม่ประสงค์ออกนาม
        cust_code = "N/A"
        cust_name = "ผู้ไม่ประสงค์ออกนาม"
    else:
        # มีข้อมูลมาบางส่วน → ใส่ให้ครบ
        cust_code = raw_code or "N/A"
        cust_name = raw_name or "ผู้ไม่ประสงค์ออกนาม"

    quote_no = _generate_quote_no(branch)
    now = _now_iso()

    # 🔥 เรียงตามคอลัมน์จริงของ DB
    header = {
        "QuoteNo": quote_no,
        "Status": payload.get("status", "draft"),
        "CustomerCode": cust_code, 
        "SalesID": employee.get("id", ""),
        "SalesName": employee.get("name", ""),
        "CreateDate": now,
        "ExpireDate": payload.get("expireDate", ""),
        "ApproveDate": now,
        "BranchCode": branch,                   # ใช้ชื่อเดียวกับ DB
        "PaymentTerm": payload.get("paymentTerm", ""),
        "CreditTerm": payload.get("creditTerm", ""),
        "ShippingMethod": payload.get("deliveryType", ""),
        "ShippingCost": payload.get("totals", {}).get("shippingRaw", 0),
        "DiscountAmount": payload.get("discount", 0),
        "SubtotalAmount": payload.get("totals", {}).get("exVat", 0),
        "TotalAmount": payload.get("totals", {}).get("grandTotal", 0),
        "NeedsTax": "Y" if payload.get("needTaxInvoice") else "N",
        "BillTaxName": payload.get("billTaxName", ""),
        "Remark": payload.get("note", ""),
        "LastUpdate": now,
        "CustomerName": cust_name, 
        "Tel": customer.get("phone", ""),
        "ShippingCustomerPay": payload.get("totals", {}).get("shippingCustomerPay", 0),
        
    }

    # 🔥 INSERT ด้วยลำดับคอลัมน์ที่ถูกต้อง 100%
    cur.execute("""
        INSERT INTO Quote_Header (
            QuoteNo, Status, CustomerCode, SalesID, SalesName,
            CreateDate, ExpireDate, ApproveDate, BranchCode,
            PaymentTerm, CreditTerm, ShippingMethod, ShippingCost,
            DiscountAmount, SubtotalAmount, TotalAmount,
            NeedsTax, BillTaxName, Remark, LastUpdate,
            CustomerName, Tel , ShippingCustomerPay
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, tuple(header.values()))


    cart = payload.get("cart", [])
    if not cart:
        raise HTTPException(400, "ต้องมีสินค้าอย่างน้อย 1 รายการ")

    lines_to_excel = []

    for item in cart:
            line = _build_line_from_payload(item)

            cur.execute("""
                INSERT INTO Quote_Line (
                    QuoteID, ItemCode, ItemName, Category,
                    Unit, Quantity, UnitPrice, TotalPrice,
                    IsGlassCut, CutInfoJson, Remark,Sqft_Sheet,VariantCode, ProductWeight
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                quote_no,
                line["ItemCode"], line["ItemName"], line["Category"],
                line["Unit"], line["Quantity"], line["UnitPrice"],
                line["TotalPrice"], line["IsGlassCut"],
                line["CutInfoJson"], line["Remark"],line["Sqft_Sheet"],line["VariantCode"],line["ProductWeight"]
            ))

            lines_to_excel.append(line)


    conn.commit()
    conn.close()

    _append_header_to_excel(header)
    _append_lines_to_excel(quote_no, lines_to_excel)

    return {"quoteNo": quote_no, "status": "created"}


# -----------------------------------------------------
# UPDATE QUOTATION
# -----------------------------------------------------
@router.put("/{quote_no:path}", summary="อัปเดตใบเสนอราคา")
def update_quotation(quote_no: str, payload: dict = Body(...)):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM Quote_Header WHERE QuoteNo=?", (quote_no,))
    original = cur.fetchone()
    if not original:
        raise HTTPException(404, f"ไม่พบใบเสนอราคา {quote_no}")

    employee = payload.get("employee") or {}
    customer = payload.get("customer") or {}
    now = _now_iso()

    # -------------------------------
    #  กำหนด default ผู้ไม่ประสงค์ออกนาม ตอนอัปเดต
    # -------------------------------
    raw_code = (customer.get("code") or "").strip()
    raw_name = (customer.get("name") or "").strip()
    raw_phone = (customer.get("phone") or "").strip()

    if not raw_code and not raw_name:
        cust_code = "N/A"
        cust_name = "ผู้ไม่ประสงค์ออกนาม"
    else:
        cust_code = raw_code or "N/A"
        cust_name = raw_name or "ผู้ไม่ประสงค์ออกนาม"

    # 🔥 จัดลำดับ header ตาม DB
    header = {
        "Status": payload.get("status", "draft"),
        "CustomerCode": cust_code,
        "SalesID": employee.get("id", ""),
        "SalesName": employee.get("name", ""),
        "ExpireDate": payload.get("expireDate", ""),
        "ApproveDate": now,
        "BranchCode": employee.get("branchId", ""),
        "PaymentTerm": payload.get("paymentTerm", ""),
        "CreditTerm": payload.get("creditTerm", ""),
        "ShippingMethod": payload.get("deliveryType", ""),
        "ShippingCost": payload.get("totals", {}).get("shippingRaw", 0),
        "DiscountAmount": payload.get("discount", 0),
        "SubtotalAmount": payload.get("totals", {}).get("exVat", 0),
        "TotalAmount": payload.get("totals", {}).get("grandTotal", 0),
        "NeedsTax": "Y" if payload.get("needTaxInvoice") else "N",
        "BillTaxName": payload.get("billTaxName", ""),
        "Remark": payload.get("note", ""),
        "LastUpdate": now,
        "CustomerName": cust_name, 
        "Tel": customer.get("phone", ""),
        "ShippingCustomerPay": payload.get("totals", {}).get("shippingCustomerPay", 0),

    }

    # 🔥 UPDATE ตามลำดับที่ถูกต้อง 100%
    cur.execute("""
        UPDATE Quote_Header SET
            Status=?, CustomerCode=?, SalesID=?, SalesName=?,
            ExpireDate=?, ApproveDate=?, BranchCode=?,
            PaymentTerm=?, CreditTerm=?, ShippingMethod=?, ShippingCost=?,
            DiscountAmount=?, SubtotalAmount=?, TotalAmount=?,
            NeedsTax=?, BillTaxName=?, Remark=?, LastUpdate=?,
            CustomerName=?, Tel=? , ShippingCustomerPay=?
        WHERE QuoteNo=?
    """, (
        header["Status"],
        header["CustomerCode"],
        header["SalesID"],
        header["SalesName"],
        header["ExpireDate"],
        header["ApproveDate"],
        header["BranchCode"],
        header["PaymentTerm"],
        header["CreditTerm"],
        header["ShippingMethod"],
        header["ShippingCost"],
        header["DiscountAmount"],
        header["SubtotalAmount"],
        header["TotalAmount"],
        header["NeedsTax"],
        header["BillTaxName"],
        header["Remark"],
        header["LastUpdate"],
        header["CustomerName"],
        header["Tel"],
        header["ShippingCustomerPay"],   # ⭐ FIX สำคัญมาก!
        quote_no                         # ⭐ ตัวสุดท้าย
    ))


    cur.execute("DELETE FROM Quote_Line WHERE QuoteID=?", (quote_no,))

    cart = payload.get("cart", [])
    if not cart:
        raise HTTPException(400, "ต้องมีสินค้าอย่างน้อย 1 รายการ")

    lines_to_excel = []
    for item in cart:
        line = _build_line_from_payload(item)

        cur.execute("""
            INSERT INTO Quote_Line (
                QuoteID, ItemCode, ItemName, Category,
                Unit, Quantity, UnitPrice, TotalPrice,
                IsGlassCut, CutInfoJson, Remark ,Sqft_Sheet,VariantCode, ProductWeight
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            quote_no,
            line["ItemCode"], line["ItemName"], line["Category"],
            line["Unit"], line["Quantity"], line["UnitPrice"],
            line["TotalPrice"], line["IsGlassCut"],
            line["CutInfoJson"], line["Remark"],line["Sqft_Sheet"],line["VariantCode"],line["ProductWeight"]
        ))

        lines_to_excel.append(line)


    conn.commit()
    conn.close()

    updated_header = dict(original)
    updated_header.update(header)
    updated_header["QuoteNo"] = quote_no
    updated_header["CreateDate"] = original["CreateDate"]

    _append_header_to_excel(updated_header)
    _append_lines_to_excel(quote_no, lines_to_excel)

    return {"quoteNo": quote_no, "status": "updated"}


# -----------------------------------------------------
# LIST / HISTORY
# -----------------------------------------------------
@router.get("", summary="โหลดรายการใบเสนอราคาแบบทั้งหมด")
def list_quotations(status: str = None):
    conn = get_conn()
    cur = conn.cursor()

    if status:
        cur.execute("""
            SELECT * FROM Quote_Header
            WHERE Status = ?
            ORDER BY LastUpdate DESC
            LIMIT 5
        """, (status,))
    else:
        cur.execute("""
            SELECT * FROM Quote_Header
            ORDER BY LastUpdate DESC
            LIMIT 5
        """)

    headers = [normalize_keys(dict(r)) for r in cur.fetchall()]
    result = []

    for h in headers:
        quote_no = h["QuoteNo"]

        cur.execute("SELECT * FROM Quote_Line WHERE QuoteID=?", (quote_no,))
        lines = [normalize_keys(dict(r)) for r in cur.fetchall()]

        result.append({
            "quoteNo": quote_no,
            "id": quote_no,
            "customer": {
                "id": h["CustomerCode"],
                "code": h["CustomerCode"],
                "name": h["CustomerName"],
                "phone": h.get("Tel", "")
            },
            "employee": {
                "id": h["SalesID"],
                "name": h["SalesName"]
            },
            "createdAt": h["CreateDate"],
            "updatedAt": h["LastUpdate"],
            "totals": {
                "grandTotal": h["TotalAmount"],
                "exVat": h["SubtotalAmount"],
                "shippingRaw": h["ShippingCost"],
            },
            "cart": [
                {
                    "sku": ln["ItemCode"],
                    "name": ln["ItemName"],
                    "qty": ln["Quantity"],
                    "price": ln["UnitPrice"],
                    "lineTotal": ln["TotalPrice"],
                    "category": ln["Category"],
                    "unit": ln["Unit"],
                    "sqft_sheet": ln.get("Sqft_Sheet", 0),
                    "product_weight": ln.get("ProductWeight", 0),
                    "variantCode": ln.get("VariantCode", ""),

                }
                for ln in lines
            ]
        })

    conn.close()
    return result


# -----------------------------------------------------
# GET SINGLE QUOTATION
# -----------------------------------------------------
@router.get("/{quote_no:path}", summary="โหลดใบเสนอราคาแบบเต็ม")
def get_quotation(quote_no: str):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM Quote_Header WHERE QuoteNo=?", (quote_no,))
    header = cur.fetchone()
    if not header:
        raise HTTPException(404, f"ไม่พบใบเสนอราคา {quote_no}")

    header = normalize_keys(dict(header))

    cur.execute("SELECT * FROM Quote_Line WHERE QuoteID=?", (quote_no,))
    lines = [normalize_keys(dict(r)) for r in cur.fetchall()]

    conn.close()

    return {
        "header": header,
        "lines": [
            {
                **ln,
                "product_weight": ln.get("ProductWeight", 0),
                "variantCode": ln.get("VariantCode", ""),
            }
            for ln in lines
            ]

    }


# -----------------------------------------------------
# DELETE
# -----------------------------------------------------
@router.delete("/{quote_no:path}", summary="ยกเลิกใบเสนอราคา")
def cancel_quotation(quote_no: str):
    conn = get_conn()
    cur = conn.cursor()

    # ตรวจสอบว่ามีใบนี้ไหม
    cur.execute("SELECT * FROM Quote_Header WHERE QuoteNo=?", (quote_no,))
    if not cur.fetchone():
        raise HTTPException(404, "ไม่พบใบเสนอราคา")

    # เปลี่ยนสถานะเป็น cancelled
    cur.execute("""
        UPDATE Quote_Header 
        SET Status = 'cancelled', LastUpdate = ?
        WHERE QuoteNo = ?
    """, (_now_iso(), quote_no))

    conn.commit()
    conn.close()

    return {"cancelled": quote_no}

